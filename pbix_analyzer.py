#!/usr/bin/env python3
"""
PBIX Analyzer v1.0
Analyzes Power BI files (.pbix) and creates an Excel report
showing where each column is used in visuals and Power Query.
"""

import sys
import subprocess
import os

# Force UTF-8 on Windows terminal (needed for emojis/accented chars)
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ============================================================
# Auto-install dependencies if needed
# ============================================================
def ensure_deps():
    missing = []
    for pkg in ['openpyxl']:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"Installing: {', '.join(missing)}...")
        subprocess.check_call(
            [sys.executable, '-m', 'pip', 'install'] + missing + ['--quiet'],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )

ensure_deps()

# ============================================================
# Imports
# ============================================================
import zipfile
import json
import re
import tempfile
import shutil
from pathlib import Path
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# File reading helpers
# ============================================================

def read_layout_json(file_path):
    """Reads the Layout file (typically UTF-16 LE)"""
    for enc in ('utf-16-le', 'utf-16', 'utf-8-sig', 'utf-8'):
        try:
            with open(file_path, 'r', encoding=enc) as f:
                return json.loads(f.read())
        except Exception:
            pass
    # Binary fallback
    with open(file_path, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        text = raw[2:].decode('utf-16-le', errors='replace')
    elif raw[:2] == b'\xfe\xff':
        text = raw[2:].decode('utf-16-be', errors='replace')
    elif raw[:3] == b'\xef\xbb\xbf':
        text = raw[3:].decode('utf-8', errors='replace')
    else:
        text = raw.decode('utf-16-le', errors='replace')
    return json.loads(text)


def read_text_file(path):
    """Reads a text file trying multiple encodings"""
    for enc in ('utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'windows-1252'):
        try:
            with open(path, 'r', encoding=enc) as f:
                return f.read()
        except Exception:
            pass
    with open(path, 'rb') as f:
        return f.read().decode('utf-8', errors='replace')


def safe_json(s, default=None):
    """Parses a JSON string without raising an exception"""
    if s is None:
        return default if default is not None else {}
    if isinstance(s, (dict, list)):
        return s
    try:
        return json.loads(s)
    except Exception:
        return default if default is not None else {}


# ============================================================
# Result classes
# ============================================================

class UsageRecord:
    """Represents a column/measure usage instance"""
    __slots__ = ('table', 'column', 'usage_type', 'page', 'visual_type', 'visual_title', 'source')

    def __init__(self, table, column, usage_type, page, visual_type, visual_title, source):
        self.table = str(table or '')
        self.column = str(column or '')
        self.usage_type = str(usage_type or '')
        self.page = str(page or '')
        self.visual_type = str(visual_type or '')
        self.visual_title = str(visual_title or '')
        self.source = str(source or '')

    @property
    def full_ref(self):
        return f"{self.table}[{self.column}]" if self.table else self.column


# ============================================================
# Main parser
# ============================================================

class PBIXParser:
    def __init__(self, pbix_path, log=None):
        self.pbix_path = Path(pbix_path)
        self.log = log or print
        self.records = []
        self.pq_queries = {}           # name -> M code
        self.model_tables = {}         # table -> {columns, measures}
        self.diagram_tables = []       # table names from DiagramLayout
        self.dax_queries = {}          # name -> DAX code
        self._record_keys = set()      # for deduplication

    # ----------------------------------------------------------
    # Main entry point
    # ----------------------------------------------------------

    def parse(self):
        tmp = tempfile.mkdtemp(prefix='pbix_')
        try:
            self._extract(tmp)
            self._parse_all(tmp)
            return True
        except Exception as e:
            self.log(f"❌ Unexpected error: {e}")
            import traceback
            self.log(traceback.format_exc())
            return False
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    # ----------------------------------------------------------
    # Extraction
    # ----------------------------------------------------------

    def _extract(self, tmp):
        self.log("📦 Extracting PBIX...")
        with zipfile.ZipFile(self.pbix_path, 'r') as z:
            z.extractall(tmp)
        files = []
        for root, _, fs in os.walk(tmp):
            for f in fs:
                files.append(os.path.relpath(os.path.join(root, f), tmp))
        self.log(f"   {len(files)} internal files found:")
        for f in sorted(files):
            self.log(f"   • {f}")

    # ----------------------------------------------------------
    # Orchestrates component parsing
    # ----------------------------------------------------------

    def _parse_all(self, tmp):
        # Detect format: old (Report/Layout) or new (Report/definition/pages/)
        layout_old = os.path.join(tmp, 'Report', 'Layout')
        layout_new = os.path.join(tmp, 'Report', 'definition', 'pages')

        if os.path.exists(layout_old):
            self.log("\n🎨 Analyzing visuals and filters (old format)...")
            self._parse_layout(layout_old)
            self.log(f"   ✅ {len(self.records)} references found")
        elif os.path.exists(layout_new):
            self.log("\n🎨 Analyzing visuals and filters (new format)...")
            self._parse_new_format(tmp, layout_new)
            self.log(f"   ✅ {len(self.records)} references found")
        else:
            self.log("⚠️  Layout file not found")

        mashup = os.path.join(tmp, 'Mashup')
        if os.path.exists(mashup):
            self.log("\n🔧 Analyzing Power Query...")
            self._parse_mashup(mashup, tmp)
            self.log(f"   ✅ {len(self.pq_queries)} PQ queries found")
        else:
            self.log("⚠️  Power Query (Mashup) not found")
            self.log("   (PBIX with embedded DataModel — Power Query is compressed in DataModel)")
            self.log("   Tip: use 'pbi-tools extract' to access the full M code")

        schema = os.path.join(tmp, 'DataModelSchema')
        if os.path.exists(schema):
            self.log("\n📊 Analyzing data model schema...")
            self._parse_schema(schema)

        diag = os.path.join(tmp, 'DiagramLayout')
        if os.path.exists(diag):
            self._parse_diagram_layout(diag)
            self.log(f"   ✅ {len(self.diagram_tables)} tables identified in the model")

        # DAX Queries (queries saved in the PBIX file)
        dax_dir = os.path.join(tmp, 'DAXQueries')
        if os.path.exists(dax_dir):
            self.log("\n📐 Analyzing saved DAX Queries...")
            self._parse_dax_queries(dax_dir)

    # ----------------------------------------------------------
    # Layout (visuals) — old format
    # ----------------------------------------------------------

    def _parse_layout(self, path):
        try:
            layout = read_layout_json(path)
        except Exception as e:
            self.log(f"   ❌ Could not read Layout: {e}")
            return

        # Report-level filters
        rfilters = safe_json(layout.get('filters', '[]'), [])
        self._parse_filters(rfilters, 'Report', 'Report Filter', '', 'Report-level Filter')

        pages = layout.get('sections', [])
        self.log(f"   Pages: {len(pages)}")

        for page in pages:
            pname = page.get('displayName') or page.get('name', '?')

            # Page-level filters
            pfilters = safe_json(page.get('filters', '[]'), [])
            self._parse_filters(pfilters, pname, 'Page Filter', '', 'Page-level Filter')

            visuals = page.get('visualContainers', [])
            self.log(f"   Page '{pname}': {len(visuals)} visuals")

            for v in visuals:
                self._parse_visual(v, pname)

    # ----------------------------------------------------------
    # New PBIX format (2024+): Report/definition/pages/
    # ----------------------------------------------------------

    def _parse_new_format(self, tmp, pages_dir):
        """Parser for the new PBIX format with individual JSON files"""
        # Read page order
        pages_meta_path = os.path.join(pages_dir, 'pages.json')
        try:
            pages_meta = json.loads(read_text_file(pages_meta_path))
            page_order = pages_meta.get('pageOrder', [])
        except Exception:
            # Fallback: list folders
            page_order = [d for d in os.listdir(pages_dir)
                          if os.path.isdir(os.path.join(pages_dir, d))]

        # Report-level filters
        report_json_path = os.path.join(tmp, 'Report', 'definition', 'report.json')
        if os.path.exists(report_json_path):
            try:
                rj = json.loads(read_text_file(report_json_path))
                fc = rj.get('filterConfig', {})
                self._parse_new_filter_config(fc, 'Report', 'Report Filter', '', 'Report-level Filter')
            except Exception:
                pass

        self.log(f"   Pages: {len(page_order)}")

        for page_id in page_order:
            page_dir = os.path.join(pages_dir, page_id)
            if not os.path.isdir(page_dir):
                continue

            # Read page info
            page_json_path = os.path.join(page_dir, 'page.json')
            pname = page_id
            try:
                pj = json.loads(read_text_file(page_json_path))
                pname = pj.get('displayName', page_id)
                # Page filters
                self._parse_new_filter_config(
                    pj.get('filterConfig', {}), pname,
                    'Page Filter', '', 'Page-level Filter')
            except Exception:
                pass

            # Read visuals
            visuals_dir = os.path.join(page_dir, 'visuals')
            if not os.path.isdir(visuals_dir):
                continue

            visual_ids = [d for d in os.listdir(visuals_dir)
                          if os.path.isdir(os.path.join(visuals_dir, d))]
            self.log(f"   Page '{pname}': {len(visual_ids)} visuals")

            for vid in visual_ids:
                vjson_path = os.path.join(visuals_dir, vid, 'visual.json')
                if not os.path.exists(vjson_path):
                    continue
                try:
                    vj = json.loads(read_text_file(vjson_path))
                    self._parse_new_visual(vj, pname)
                except Exception as e:
                    self.log(f"   ⚠️ Visual {vid}: {e}")

    def _parse_new_visual(self, vj, page):
        """Parses a visual in the new format"""
        vis = vj.get('visual', {})
        vtype = vis.get('visualType', 'unknown')
        vtitle = self._get_title_new(vis)

        query_state = vis.get('query', {}).get('queryState', {})

        for role_name, role_data in query_state.items():
            for proj in role_data.get('projections', []):
                result = self._resolve_new_projection(proj, role_name)
                if result:
                    table, col, utype = result
                    self._add(table, col, utype, page, vtype, vtitle, 'Visual - Field')

        # Visual filters
        fc = vis.get('filterConfig', {})
        self._parse_new_filter_config(fc, page, vtype, vtitle, 'Visual Filter')

    def _resolve_new_projection(self, proj, role_name):
        """Resolves a projection in the new format -> (table, column, type)"""
        field = proj.get('field', {})

        if 'Column' in field:
            c = field['Column']
            entity = c.get('Expression', {}).get('SourceRef', {}).get('Entity', '')
            prop = c.get('Property', '')
            if prop:
                return (entity, prop, f'Column ({role_name})')

        if 'Measure' in field:
            m = field['Measure']
            entity = m.get('Expression', {}).get('SourceRef', {}).get('Entity', '')
            prop = m.get('Property', '')
            if prop:
                return (entity, prop, f'Measure ({role_name})')

        if 'HierarchyLevel' in field:
            hl = field['HierarchyLevel']
            h = hl.get('Expression', {}).get('Hierarchy', {})
            entity = h.get('Expression', {}).get('SourceRef', {}).get('Entity', '')
            hier = h.get('Hierarchy', '')
            level = hl.get('Level', '')
            col = f"{hier} > {level}" if hier else level
            if col:
                return (entity, col, f'Hierarchy ({role_name})')

        # Fallback: queryRef = "Table.Column"
        qref = proj.get('queryRef', '')
        if qref and '.' in qref:
            parts = qref.split('.', 1)
            return (parts[0], parts[1], f'Field ({role_name})')

        return None

    def _parse_new_filter_config(self, fc, page, vtype, vtitle, source):
        """Parses filterConfig in the new format"""
        if not isinstance(fc, dict):
            return
        for f in fc.get('filters', []):
            field = f.get('field', {})
            for key in ('Column', 'Measure'):
                if key in field:
                    item = field[key]
                    entity = item.get('Expression', {}).get('SourceRef', {}).get('Entity', '')
                    prop = item.get('Property', '')
                    if prop:
                        self._add(entity, prop, 'Filter', page, vtype, vtitle, source)

    def _get_title_new(self, vis):
        """Extracts title in the new format"""
        try:
            objects = vis.get('objects', {}) or vis.get('visualContainerObjects', {})
            for key in ('title', 'visualTitle'):
                arr = objects.get(key, [])
                if arr:
                    p = arr[0].get('properties', {})
                    t = p.get('text', {})
                    if isinstance(t, dict):
                        lit = t.get('expr', {}).get('Literal', {})
                        val = lit.get('Value', '')
                        if val:
                            return val.strip("'\"")
        except Exception:
            pass
        return ''

    # ----------------------------------------------------------
    # Old format: parse_visual
    # ----------------------------------------------------------

    def _parse_visual(self, v, page):
        config = safe_json(v.get('config', '{}'))
        query = safe_json(v.get('query', '{}'))
        dtrans = safe_json(v.get('dataTransforms', '{}'))
        filters = safe_json(v.get('filters', '[]'), [])

        sv = config.get('singleVisual', {})
        vtype = sv.get('visualType', 'unknown')
        vtitle = self._get_title(config)

        alias_map = self._build_alias_map(query)

        # Fields in query Select
        for cmd in query.get('Commands', []):
            q = cmd.get('SemanticQueryDataShapeCommand', {}).get('Query', {})
            for sel in q.get('Select', []):
                r = self._resolve_select(sel, alias_map)
                if r:
                    table, col, utype = r
                    self._add(table, col, utype, page, vtype, vtitle, 'Visual - Field')

            # Where clause (internal query filters)
            for where in q.get('Where', []):
                for table, col in self._cols_from_expr(where, alias_map):
                    self._add(table, col, 'Query Filter', page, vtype, vtitle, 'Visual - Query Filter')

        # dataTransforms (role info: Axis, Value, Legend, etc.)
        for sel in dtrans.get('selects', []):
            r = self._resolve_dt_select(sel)
            if r:
                table, col, role = r
                self._add(table, col, f'Field ({role})', page, vtype, vtitle, 'Visual - Role')

        # Visual filters
        self._parse_filters(filters, page, vtype, vtitle, 'Visual Filter')

    def _get_title(self, config):
        try:
            sv = config.get('singleVisual', {})
            for src in [sv.get('vcObjects', {}), sv.get('objects', {})]:
                for key in ('title', 'visualTitle'):
                    arr = src.get(key, [])
                    if arr:
                        p = arr[0].get('properties', {})
                        t = p.get('text', {})
                        if isinstance(t, dict):
                            lit = t.get('expr', {}).get('Literal', {})
                            val = lit.get('Value', '')
                            if val:
                                return val.strip("'\"")
                            val2 = t.get('value', '')
                            if val2:
                                return str(val2)
        except Exception:
            pass
        return ''

    def _build_alias_map(self, query):
        m = {}
        for cmd in query.get('Commands', []):
            q = cmd.get('SemanticQueryDataShapeCommand', {}).get('Query', {})
            for f in q.get('From', []):
                if f.get('Name') and f.get('Entity'):
                    m[f['Name']] = f['Entity']
        return m

    def _resolve_select(self, sel, alias_map):
        """Resolves a Select item -> (table, column, type)"""
        func_map = {0: 'Sum', 1: 'Average', 2: 'Count', 3: 'Min', 4: 'Max', 5: 'CountRows'}

        if 'Column' in sel:
            c = sel['Column']
            src = c.get('Expression', {}).get('SourceRef', {}).get('Source', '')
            prop = c.get('Property', '')
            if prop:
                return (alias_map.get(src, src), prop, 'Column')

        if 'Measure' in sel:
            m = sel['Measure']
            src = m.get('Expression', {}).get('SourceRef', {}).get('Source', '')
            prop = m.get('Property', '')
            if prop:
                return (alias_map.get(src, src), prop, 'Measure')

        if 'Aggregation' in sel:
            agg = sel['Aggregation']
            func = func_map.get(agg.get('Function', 0), 'Aggregation')
            expr = agg.get('Expression', {})
            if 'Column' in expr:
                c = expr['Column']
                src = c.get('Expression', {}).get('SourceRef', {}).get('Source', '')
                prop = c.get('Property', '')
                if prop:
                    return (alias_map.get(src, src), prop, f'Aggregation ({func})')

        if 'HierarchyLevel' in sel:
            hl = sel['HierarchyLevel']
            h = hl.get('Expression', {}).get('Hierarchy', {})
            src = h.get('Expression', {}).get('SourceRef', {}).get('Source', '')
            hier = h.get('Hierarchy', '')
            level = hl.get('Level', '')
            col = f"{hier} > {level}" if hier else level
            if col:
                return (alias_map.get(src, src), col, 'Hierarchy')

        # Fallback by Name field: "Table.Column"
        name = sel.get('Name', '')
        if name:
            m2 = re.search(r'([A-Za-z_\u00C0-\u024F][\w\s\u00C0-\u024F]*)\.([A-Za-z_\u00C0-\u024F][\w\s\u00C0-\u024F]*)', name)
            if m2:
                return (m2.group(1).strip(), m2.group(2).strip(), 'Field')

        return None

    def _resolve_dt_select(self, sel):
        """Resolves a dataTransforms item -> (table, column, role)"""
        qref = sel.get('queryRef', '')
        if qref and '.' in qref:
            parts = qref.split('.', 1)
            roles = sel.get('roles', [])
            role = ', '.join(roles) if roles else 'Field'
            return (parts[0], parts[1], role)

        if 'Column' in sel:
            c = sel['Column']
            entity = c.get('Expression', {}).get('SourceRef', {}).get('Entity', '')
            prop = c.get('Property', '')
            if prop:
                roles = sel.get('roles', [])
                return (entity, prop, ', '.join(roles) if roles else 'Column')
        return None

    def _cols_from_expr(self, obj, alias_map):
        """Recursively extracts (table, column) pairs from an expression"""
        result = []
        if not isinstance(obj, dict):
            return result

        for key in ('Column', 'Measure'):
            if key in obj:
                item = obj[key]
                src_ref = item.get('Expression', {}).get('SourceRef', {})
                src = src_ref.get('Source') or src_ref.get('Entity', '')
                prop = item.get('Property', '')
                if prop:
                    result.append((alias_map.get(src, src), prop))

        for v in obj.values():
            if isinstance(v, dict):
                result.extend(self._cols_from_expr(v, alias_map))
            elif isinstance(v, list):
                for item in v:
                    if isinstance(item, dict):
                        result.extend(self._cols_from_expr(item, alias_map))
        return result

    def _parse_filters(self, filters, page, vtype, vtitle, source):
        if not isinstance(filters, list):
            return
        for f in filters:
            if not isinstance(f, dict):
                continue
            # Internal format: whereItems
            for wi in f.get('whereItems', []):
                for table, col in self._cols_from_expr(wi.get('condition', {}), {}):
                    self._add(table, col, 'Filter', page, vtype, vtitle, source)
            # API format: target
            tgt = f.get('target', {})
            if isinstance(tgt, dict):
                tbl = tgt.get('table', '')
                col = tgt.get('column', '') or tgt.get('measure', '')
                if col:
                    self._add(tbl, col, 'Filter', page, vtype, vtitle, source)

    def _add(self, table, column, usage_type, page, vtype, vtitle, source):
        if not column:
            return
        # Deduplicate: same column+usage+page+visual combination
        key = (table, column, usage_type, page, vtype, vtitle)
        if key in self._record_keys:
            return
        self._record_keys.add(key)
        self.records.append(UsageRecord(table, column, usage_type, page, vtype, vtitle, source))

    # ----------------------------------------------------------
    # Mashup (Power Query)
    # ----------------------------------------------------------

    def _parse_mashup(self, mashup_path, tmp):
        mdir = os.path.join(tmp, '_mdir')
        os.makedirs(mdir, exist_ok=True)
        try:
            with zipfile.ZipFile(mashup_path, 'r') as z:
                z.extractall(mdir)
            sec = os.path.join(mdir, 'Formulas', 'Section1.m')
            if os.path.exists(sec):
                content = read_text_file(sec)
                self._parse_section_m(content)
        except Exception as e:
            self.log(f"   ⚠️ Mashup error: {e}")
        finally:
            shutil.rmtree(mdir, ignore_errors=True)

    def _parse_section_m(self, content):
        # Remove section header
        content = re.sub(r'^\s*section\s+\S+\s*;', '', content, flags=re.MULTILINE)
        # Split by "shared" declarations
        parts = re.split(r'\n(?=\s*shared\s)', content)
        for part in parts:
            part = part.strip()
            if not part:
                continue
            m = re.match(r'\s*shared\s+(#?"[^"]*"|[^\s=]+)\s*=', part)
            if m:
                raw_name = m.group(1)
                name = raw_name.strip().strip('"').lstrip('#').strip('"')
                self.pq_queries[name] = part

    # ----------------------------------------------------------
    # DiagramLayout (extracts table names)
    # ----------------------------------------------------------

    def _parse_diagram_layout(self, diag_path):
        try:
            content = read_text_file(diag_path)
            diag = json.loads(content)
            tables = set()
            for diagram in diag.get('diagrams', []):
                for node in diagram.get('nodes', []):
                    t = node.get('nodeIndex', '')
                    if t:
                        tables.add(t)
            self.diagram_tables = sorted(tables)
        except Exception as e:
            self.log(f"   ⚠️ DiagramLayout: {e}")

    # ----------------------------------------------------------
    # DAX Queries (saved inside the PBIX)
    # ----------------------------------------------------------

    def _parse_dax_queries(self, dax_dir):
        """Reads DAX queries saved in the PBIX (DAXQueries folder)"""
        import urllib.parse
        count = 0
        for root, _, files in os.walk(dax_dir):
            for f in files:
                if f.endswith('.dax'):
                    fpath = os.path.join(root, f)
                    name = urllib.parse.unquote(f.replace('.dax', ''))
                    try:
                        # DAX files are UTF-16 LE (same as Layout)
                        code = read_layout_json.__code__  # just to ensure import
                        code = read_text_file(fpath)
                        # Clean up null bytes that appear in UTF-16 read as UTF-8
                        if '\x00' in code:
                            code = code.replace('\x00', '')
                        self.dax_queries[name] = code
                        count += 1
                    except Exception:
                        pass
        if count:
            self.log(f"   ✅ {count} DAX queries found")

    # ----------------------------------------------------------
    # DataModelSchema
    # ----------------------------------------------------------

    def _parse_schema(self, schema_path):
        try:
            content = read_text_file(schema_path)
            schema = json.loads(content)
            model = schema.get('model', schema)
            for tbl in model.get('tables', []):
                tname = tbl.get('name', '')
                self.model_tables[tname] = {
                    'columns': [
                        {'name': c.get('name', ''), 'dataType': c.get('dataType', ''),
                         'isHidden': c.get('isHidden', False), 'expression': c.get('expression', '')}
                        for c in tbl.get('columns', [])
                    ],
                    'measures': [
                        {'name': m.get('name', ''), 'expression': m.get('expression', ''),
                         'isHidden': m.get('isHidden', False)}
                        for m in tbl.get('measures', [])
                    ]
                }
            self.log(f"   ✅ {len(self.model_tables)} tables in the model")
        except Exception as e:
            self.log(f"   ⚠️ Schema error: {e}")

    # ----------------------------------------------------------
    # PQ References
    # ----------------------------------------------------------

    def get_pq_refs(self):
        """Extracts column references from Power Query code"""
        refs = []
        seen = set()

        for qname, code in self.pq_queries.items():
            found = {}  # col -> context

            # [ColumnName]
            for col in re.findall(r'\[([^\[\]\n]{1,100})\]', code):
                col = col.strip()
                if col and not col.isdigit() and not col.startswith('#') and len(col) > 1:
                    found.setdefault(col, '[Column] Reference')

            # Table.* functions
            funcs = {
                'Table.SelectColumns': r'Table\.SelectColumns\s*\([^,]+,\s*(\{[^}]*\})',
                'Table.RemoveColumns': r'Table\.RemoveColumns\s*\([^,]+,\s*(\{[^}]*\})',
                'Table.RenameColumns': r'Table\.RenameColumns\s*\([^,]+,\s*(\{[^}]*\})',
                'Table.TransformColumns': r'Table\.TransformColumns\s*\([^,]+,\s*(\{[^}]*\})',
                'Table.ReorderColumns': r'Table\.ReorderColumns\s*\([^,]+,\s*(\{[^}]*\})',
                'Table.SplitColumn': r'Table\.SplitColumn\s*\([^,]+,\s*"([^"]+)"',
                'Table.ExpandTableColumn': r'Table\.ExpandTableColumn\s*\([^,]+,\s*"[^"]+",\s*(\{[^}]*\})',
            }
            for func, pat in funcs.items():
                for m in re.findall(pat, code, re.DOTALL):
                    for col in re.findall(r'"([^"]+)"', m):
                        found[col] = func

            for col, ctx in found.items():
                key = (qname, col, ctx)
                if key not in seen:
                    seen.add(key)
                    refs.append({'query': qname, 'column': col, 'context': ctx})

        return refs


# ============================================================
# Excel Generator
# ============================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
ZEBRA_FILL = PatternFill(start_color="DEE9F5", end_color="DEE9F5", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
CRIT_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def _style_headers(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
    ws.row_dimensions[row].height = 28


def _auto_width(ws, max_w=55):
    for col in ws.columns:
        maxlen = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                maxlen = max(maxlen, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(maxlen + 2, 8)


def _zebra(ws, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start), start=start):
        if i % 2 == 0:
            for c in row:
                if c.fill.fgColor.rgb in ('00000000', 'FFFFFFFF', ''):
                    c.fill = ZEBRA_FILL


def _clean(val, max_len=32000):
    """Cleans a string for Excel cell use (removes illegal characters)"""
    if val is None:
        return ''
    s = str(val)
    # Remove null bytes and other control chars (keep tab, newline, CR)
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
    return s[:max_len]


def create_excel(parser, output_path, log=None):
    log = log or print
    pq_refs = parser.get_pq_refs()

    # Sets for cross-referencing
    pq_cols_lower = {r['column'].lower() for r in pq_refs}
    visual_col_keys = set()
    for rec in parser.records:
        visual_col_keys.add(rec.column.lower())

    wb = openpyxl.Workbook()

    # ===========================================================
    # SHEET 1 — Column Impact (primary sheet)
    # ===========================================================
    log("   📄 Sheet: Column Impact...")
    ws1 = wb.active
    ws1.title = "⚡ Column Impact"

    headers1 = [
        'Column', 'Table',
        'Visuals\n(count)', 'Pages',
        'Visual Types', 'Usage Type',
        'In Power Query?', 'PQ Queries',
        'RISK'
    ]
    _style_headers(ws1, headers1)

    # Aggregate by column + table
    col_agg = defaultdict(lambda: {
        'visual_count': 0, 'pages': set(), 'vtypes': set(),
        'utypes': set(), 'table': '', 'pq_queries': set()
    })
    for rec in parser.records:
        k = (rec.table, rec.column)
        col_agg[k]['visual_count'] += 1
        col_agg[k]['pages'].add(rec.page)
        col_agg[k]['vtypes'].add(rec.visual_type)
        col_agg[k]['utypes'].add(rec.usage_type)
        col_agg[k]['table'] = rec.table

    # PQ cross-reference
    pq_by_col = defaultdict(set)
    for r in pq_refs:
        pq_by_col[r['column'].lower()].add(r['query'])

    row = 2
    for (table, col), data in sorted(col_agg.items(), key=lambda x: (x[0][0], x[0][1])):
        pq_qs = pq_by_col.get(col.lower(), set())
        in_pq = 'YES' if pq_qs else 'No'
        utypes = ', '.join(t for t in sorted(data['utypes']) if t)
        pages = ', '.join(sorted(data['pages']))
        vtypes = ', '.join(sorted(data['vtypes']))
        pq_str = ', '.join(sorted(pq_qs)) if pq_qs else ''

        # Risk: HIGH if used in PQ and in multiple visuals
        vc = data['visual_count']
        if pq_qs and vc > 0:
            risk = 'HIGH — PQ + Visuals'
            rfill = CRIT_FILL
        elif vc > 3:
            risk = 'MEDIUM — Multiple Visuals'
            rfill = WARN_FILL
        elif vc > 0:
            risk = 'LOW'
            rfill = GREEN_FILL
        else:
            risk = 'PQ Only'
            rfill = ZEBRA_FILL

        cells = [col, table, vc, pages, vtypes, utypes, in_pq, pq_str, risk]
        for ci, val in enumerate(cells, 1):
            c = ws1.cell(row=row, column=ci, value=val)
            if ci == 9:  # RISK column
                c.fill = rfill
                c.font = Font(bold=True)
        row += 1

    ws1.freeze_panes = 'A2'
    _auto_width(ws1)

    # ===========================================================
    # SHEET 2 — Visual Details
    # ===========================================================
    log("   📄 Sheet: Visual Details...")
    ws2 = wb.create_sheet("Visuals - Details")
    headers2 = ['Column', 'Table', 'Usage Type', 'Page', 'Visual Type', 'Visual Title', 'Source']
    _style_headers(ws2, headers2)

    row = 2
    for rec in sorted(parser.records, key=lambda r: (r.column, r.page, r.visual_type)):
        ws2.cell(row=row, column=1, value=rec.column)
        ws2.cell(row=row, column=2, value=rec.table)
        ws2.cell(row=row, column=3, value=rec.usage_type)
        ws2.cell(row=row, column=4, value=rec.page)
        ws2.cell(row=row, column=5, value=rec.visual_type)
        ws2.cell(row=row, column=6, value=rec.visual_title)
        ws2.cell(row=row, column=7, value=rec.source)
        row += 1

    _zebra(ws2)
    ws2.freeze_panes = 'A2'
    _auto_width(ws2)

    # ===========================================================
    # SHEET 3 — Power Query - Referenced Columns
    # ===========================================================
    log("   📄 Sheet: Power Query - Columns...")
    ws3 = wb.create_sheet("Power Query - Columns")
    headers3 = ['Referenced Column', 'Power Query Name', 'Context / M Function']
    _style_headers(ws3, headers3)

    row = 2
    for r in sorted(pq_refs, key=lambda x: (x['column'], x['query'])):
        ws3.cell(row=row, column=1, value=r['column'])
        ws3.cell(row=row, column=2, value=r['query'])
        ws3.cell(row=row, column=3, value=r['context'])
        row += 1

    _zebra(ws3)
    ws3.freeze_panes = 'A2'
    _auto_width(ws3)

    # ===========================================================
    # SHEET 4 — Power Query Code (M)
    # ===========================================================
    log("   📄 Sheet: Power Query Code...")
    ws4 = wb.create_sheet("Power Query Code")
    _style_headers(ws4, ['Query / Table', 'M Code'])
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 120

    row = 2
    for qname, code in sorted(parser.pq_queries.items()):
        ws4.cell(row=row, column=1, value=qname)
        cell = ws4.cell(row=row, column=2, value=_clean(code))
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        lines = code.count('\n') + 1 if code else 1
        ws4.row_dimensions[row].height = min(lines * 12, 300)
        row += 1

    ws4.freeze_panes = 'A2'

    # ===========================================================
    # SHEET 5 — Data Model (if available)
    # ===========================================================
    if parser.model_tables:
        log("   📄 Sheet: Data Model...")
        ws5 = wb.create_sheet("Data Model")
        headers5 = ['Table', 'Name', 'Type', 'Data Type', 'Hidden', 'DAX Expression']
        _style_headers(ws5, headers5)

        row = 2
        for tname in sorted(parser.model_tables):
            tdata = parser.model_tables[tname]
            for col in sorted(tdata['columns'], key=lambda c: c['name']):
                ws5.cell(row=row, column=1, value=tname)
                ws5.cell(row=row, column=2, value=col['name'])
                ws5.cell(row=row, column=3, value='Column')
                ws5.cell(row=row, column=4, value=col['dataType'])
                ws5.cell(row=row, column=5, value='Yes' if col['isHidden'] else 'No')
                expr = col.get('expression', '') or ''
                ws5.cell(row=row, column=6, value=expr[:500])
                row += 1
            for msr in sorted(tdata['measures'], key=lambda m: m['name']):
                ws5.cell(row=row, column=1, value=tname)
                ws5.cell(row=row, column=2, value=msr['name'])
                ws5.cell(row=row, column=3, value='Measure')
                ws5.cell(row=row, column=5, value='Yes' if msr['isHidden'] else 'No')
                expr = msr.get('expression', '') or ''
                ws5.cell(row=row, column=6, value=expr[:500])
                row += 1

        _zebra(ws5)
        ws5.freeze_panes = 'A2'
        _auto_width(ws5)

    # ===========================================================
    # SHEET 6 — Model Tables (from DiagramLayout)
    # ===========================================================
    if parser.diagram_tables and not parser.model_tables:
        log("   📄 Sheet: Model Tables...")
        ws6 = wb.create_sheet("Model Tables")
        _style_headers(ws6, ['Table (internal name)', 'Used in Visuals?', 'Columns Referenced in Visuals'])
        visual_tables = {r.table.lower() for r in parser.records if r.table}
        row = 2
        for tname in parser.diagram_tables:
            used = 'Yes' if tname.lower() in visual_tables else 'No'
            cols_used = sorted({r.column for r in parser.records if r.table.lower() == tname.lower()})
            ws6.cell(row=row, column=1, value=tname)
            ws6.cell(row=row, column=2, value=used)
            ws6.cell(row=row, column=3, value=', '.join(cols_used))
            row += 1
        _zebra(ws6)
        ws6.freeze_panes = 'A2'
        _auto_width(ws6)

    # ===========================================================
    # SHEET 7 — DAX Queries (if any)
    # ===========================================================
    if parser.dax_queries:
        log("   📄 Sheet: DAX Queries...")
        ws7 = wb.create_sheet("DAX Queries")
        _style_headers(ws7, ['Query Name', 'DAX Code'])
        ws7.column_dimensions['A'].width = 35
        ws7.column_dimensions['B'].width = 120
        row = 2
        for qname, code in sorted(parser.dax_queries.items()):
            ws7.cell(row=row, column=1, value=qname)
            cell = ws7.cell(row=row, column=2, value=_clean(code))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            lines = code.count('\n') + 1 if code else 1
            ws7.row_dimensions[row].height = min(lines * 13, 300)
            row += 1
        ws7.freeze_panes = 'A2'

    # ===========================================================
    # Save
    # ===========================================================
    wb.save(output_path)
    log(f"   ✅ Excel saved: {output_path}")


# ============================================================
# Graphical Interface (GUI)
# ============================================================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("PBIX Analyzer — Column Impact")
        self.root.geometry("720x580")
        self.root.minsize(600, 480)
        self.root.configure(bg='#F0F4FA')

        self._output_path = None
        self._pbix = tk.StringVar()
        self._out = tk.StringVar()

        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self.root, bg='#1F4E79', pady=12)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="Column Impact Analyzer — Power BI",
                 font=('Segoe UI', 14, 'bold'), bg='#1F4E79', fg='white').pack()
        tk.Label(hdr, text='"What will break if this column changes?"',
                 font=('Segoe UI', 9, 'italic'), bg='#1F4E79', fg='#AECDE8').pack()

        # Body
        body = tk.Frame(self.root, bg='#F0F4FA', padx=20, pady=12)
        body.pack(fill=tk.BOTH, expand=True)

        # PBIX file row
        tk.Label(body, text="PBIX File:", font=('Segoe UI', 10, 'bold'),
                 bg='#F0F4FA').grid(row=0, column=0, sticky='w', pady=(0, 2))

        fr1 = tk.Frame(body, bg='#F0F4FA')
        fr1.grid(row=1, column=0, sticky='ew', pady=(0, 10))
        body.columnconfigure(0, weight=1)

        pbix_e = tk.Entry(fr1, textvariable=self._pbix, font=('Segoe UI', 10))
        pbix_e.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(fr1, text="📂  Browse PBIX...", command=self._browse_pbix,
                  font=('Segoe UI', 9), bg='#2E75B6', fg='white',
                  relief='flat', padx=10, cursor='hand2').pack(side=tk.RIGHT, padx=(6, 0))

        # Output row
        tk.Label(body, text="Output Excel file:", font=('Segoe UI', 10, 'bold'),
                 bg='#F0F4FA').grid(row=2, column=0, sticky='w', pady=(0, 2))

        fr2 = tk.Frame(body, bg='#F0F4FA')
        fr2.grid(row=3, column=0, sticky='ew', pady=(0, 12))

        tk.Entry(fr2, textvariable=self._out, font=('Segoe UI', 10)).pack(
            side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(fr2, text="💾  Save as...", command=self._browse_out,
                  font=('Segoe UI', 9), bg='#2E75B6', fg='white',
                  relief='flat', padx=10, cursor='hand2').pack(side=tk.RIGHT, padx=(6, 0))

        # Action buttons
        btn_row = tk.Frame(body, bg='#F0F4FA')
        btn_row.grid(row=4, column=0, sticky='w', pady=(0, 8))

        self._btn_run = tk.Button(
            btn_row, text="🔍  ANALYZE NOW",
            command=self._run,
            font=('Segoe UI', 11, 'bold'), bg='#1F4E79', fg='white',
            relief='flat', padx=18, pady=8, cursor='hand2'
        )
        self._btn_run.pack(side=tk.LEFT)

        self._btn_open = tk.Button(
            btn_row, text="📊  Open Excel",
            command=self._open_excel,
            font=('Segoe UI', 11, 'bold'), bg='#70AD47', fg='white',
            relief='flat', padx=18, pady=8, cursor='hand2', state=tk.DISABLED
        )
        self._btn_open.pack(side=tk.LEFT, padx=(10, 0))

        # Progress bar
        self._progress = ttk.Progressbar(body, mode='indeterminate')
        self._progress.grid(row=5, column=0, sticky='ew', pady=(0, 8))

        # Log
        tk.Label(body, text="Progress:", font=('Segoe UI', 10, 'bold'),
                 bg='#F0F4FA').grid(row=6, column=0, sticky='w')

        log_frame = tk.Frame(body, bg='#F0F4FA')
        log_frame.grid(row=7, column=0, sticky='nsew', pady=(2, 0))
        body.rowconfigure(7, weight=1)

        self._log_box = tk.Text(log_frame, font=('Consolas', 9),
                                bg='#1E1E1E', fg='#D4D4D4', wrap=tk.WORD)
        sb = tk.Scrollbar(log_frame, command=self._log_box.yview)
        self._log_box.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._log_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._log("Ready. Select a PBIX file and click 'ANALYZE NOW'.")

    def _log(self, msg):
        self._log_box.insert(tk.END, msg + '\n')
        self._log_box.see(tk.END)
        self.root.update_idletasks()

    def _browse_pbix(self):
        path = filedialog.askopenfilename(
            title="Select PBIX file",
            filetypes=[("Power BI", "*.pbix"), ("All files", "*.*")]
        )
        if path:
            self._pbix.set(path)
            stem = Path(path).stem
            out = Path(path).parent / f"{stem}_column_analysis.xlsx"
            self._out.set(str(out))

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="Save Excel as",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self._out.set(path)

    def _run(self):
        pbix = self._pbix.get().strip()
        if not pbix:
            messagebox.showwarning("Warning", "Please select a PBIX file first.")
            return
        if not Path(pbix).exists():
            messagebox.showerror("Error", f"File not found:\n{pbix}")
            return

        out = self._out.get().strip()
        if not out:
            out = str(Path(pbix).parent / (Path(pbix).stem + '_column_analysis.xlsx'))
            self._out.set(out)

        self._btn_run.config(state=tk.DISABLED)
        self._btn_open.config(state=tk.DISABLED)
        self._progress.start(8)

        def _worker():
            try:
                self._log(f"\n{'='*55}")
                self._log(f"Analyzing: {Path(pbix).name}")
                self._log(f"{'='*55}")

                parser = PBIXParser(pbix, self._log)
                ok = parser.parse()

                if ok:
                    self._log("\n📊 Generating Excel...")
                    create_excel(parser, out, self._log)
                    self._output_path = out
                    self._log(f"\n✅ DONE!")
                    self._log(f"File saved to:\n{out}")
                    self._btn_open.config(state=tk.NORMAL)
                    messagebox.showinfo("Done ✅",
                                        f"Analysis complete!\n\nExcel file saved to:\n{out}")
                else:
                    messagebox.showerror("Error", "Analysis failed. Check the log.")
            except Exception as e:
                self._log(f"\n❌ Error: {e}")
                import traceback
                self._log(traceback.format_exc())
                messagebox.showerror("Error", str(e))
            finally:
                self._progress.stop()
                self._btn_run.config(state=tk.NORMAL)

        threading.Thread(target=_worker, daemon=True).start()

    def _open_excel(self):
        if self._output_path and Path(self._output_path).exists():
            os.startfile(self._output_path)


# ============================================================
# Main
# ============================================================

def main():
    root = tk.Tk()
    try:
        # Modern icon (blank to use default)
        root.iconbitmap(default='')
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
